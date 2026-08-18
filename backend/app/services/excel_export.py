from __future__ import annotations

import csv
import io
import json
import logging
from datetime import datetime

from sqlalchemy.orm import Session

from app.models import Application, CandidateProfile, Job

logger = logging.getLogger(__name__)


def export_jobs_to_xlsx(db: Session) -> bytes:
    """Export jobs, candidate profile, and application tracker to XLSX."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # --- Sheet 1: Top Jobs ---
    ws_jobs = wb.active
    ws_jobs.title = "Top Jobs"

    headers = [
        "Rank", "Job Title", "Company", "Location", "Remote Type",
        "Posted Time", "Job URL", "Required Experience", "Resume Experience",
        "Match Score", "Interview Probability", "Matching Skills",
        "Missing Skills", "Strongest Resume Evidence", "Why Good Match",
        "Priority", "Recommended Action", "Status",
    ]

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col, header in enumerate(headers, 1):
        cell = ws_jobs.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    jobs = (
        db.query(Job)
        .filter(Job.match_score.isnot(None))
        .order_by(Job.match_score.desc())
        .all()
    )

    for rank, job in enumerate(jobs, 1):
        row = rank + 1
        mandatory_gaps = json.loads(job.mandatory_gaps) if job.mandatory_gaps else []
        nice_gaps = json.loads(job.nice_to_have_gaps) if job.nice_to_have_gaps else []

        matching = _get_matching_skills(job)
        missing = mandatory_gaps + nice_gaps

        values = [
            rank,
            job.title,
            job.company,
            job.location,
            job.remote_type,
            job.posted_at.strftime("%Y-%m-%d %H:%M") if job.posted_at else "Unknown",
            job.url,
            "",
            "",
            job.match_score,
            job.interview_probability or "",
            ", ".join(matching),
            ", ".join(missing),
            _get_strongest_evidence(job),
            job.match_reason or "",
            _get_priority(job),
            job.recommendation or "",
            job.status,
        ]

        for col, value in enumerate(values, 1):
            cell = ws_jobs.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Make URL clickable
        url_cell = ws_jobs.cell(row=row, column=7)
        url_cell.hyperlink = job.url
        url_cell.font = Font(color="0563C1", underline="single")

    # Auto-width columns
    for col in range(1, len(headers) + 1):
        ws_jobs.column_dimensions[get_column_letter(col)].width = max(
            12, min(40, max(len(str(ws_jobs.cell(row=r, column=col).value or "")) for r in range(1, min(20, ws_jobs.max_row + 1))))
        )

    ws_jobs.auto_filter.ref = ws_jobs.dimensions
    ws_jobs.freeze_panes = "A2"

    # --- Sheet 2: Candidate Profile ---
    ws_profile = wb.create_sheet("Candidate Profile")
    profile = db.query(CandidateProfile).first()
    if profile:
        ws_profile.cell(row=1, column=1, value="Field").font = Font(bold=True)
        ws_profile.cell(row=1, column=2, value="Value").font = Font(bold=True)

        fields = [
            ("Full Name", profile.full_name),
            ("Email", profile.email),
            ("Phone", profile.phone),
            ("Location", profile.location),
            ("Current Role", profile.current_role),
            ("Experience Years", profile.experience_years),
            ("Summary", profile.summary),
            ("Technologies", _fmt_list(profile.technologies)),
            ("Certifications", _fmt_list(profile.certifications)),
            ("Education", _fmt_list(profile.education)),
        ]
        for i, (field, value) in enumerate(fields, 2):
            ws_profile.cell(row=i, column=1, value=field)
            ws_profile.cell(row=i, column=2, value=str(value) if value else "")

    # --- Sheet 3: Skill Match ---
    ws_skills = wb.create_sheet("Skill Match")
    skill_headers = [
        "Company", "Job Title", "AWS", "Kubernetes", "Terraform",
        "CI/CD", "DevSecOps", "Python", "GitOps",
    ]
    for col, h in enumerate(skill_headers, 1):
        cell = ws_skills.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, job in enumerate(jobs, 2):
        ws_skills.cell(row=row_idx, column=1, value=job.company)
        ws_skills.cell(row=row_idx, column=2, value=job.title)
        ws_skills.cell(row=row_idx, column=3, value=job.aws_match or 0)
        ws_skills.cell(row=row_idx, column=4, value=job.kubernetes_match or 0)
        ws_skills.cell(row=row_idx, column=5, value=job.terraform_match or 0)
        ws_skills.cell(row=row_idx, column=6, value=job.cicd_match or 0)
        ws_skills.cell(row=row_idx, column=7, value=job.devsecops_match or 0)
        ws_skills.cell(row=row_idx, column=8, value=job.python_match or 0)
        ws_skills.cell(row=row_idx, column=9, value=job.gitops_match or 0)

    # --- Sheet 4: Application Tracker ---
    ws_tracker = wb.create_sheet("Application Tracker")
    tracker_headers = [
        "Company", "Job Title", "Application URL", "Priority",
        "Status", "Date Applied", "Follow-up Date", "Notes",
    ]
    for col, h in enumerate(tracker_headers, 1):
        cell = ws_tracker.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    applications = db.query(Application).all()
    for row_idx, app in enumerate(applications, 2):
        job = app.job
        ws_tracker.cell(row=row_idx, column=1, value=job.company if job else "")
        ws_tracker.cell(row=row_idx, column=2, value=job.title if job else "")
        ws_tracker.cell(row=row_idx, column=3, value=app.application_url)
        ws_tracker.cell(row=row_idx, column=4, value=_get_priority(job) if job else "")
        ws_tracker.cell(row=row_idx, column=5, value=app.status)
        ws_tracker.cell(
            row=row_idx, column=6,
            value=app.applied_at.strftime("%Y-%m-%d") if app.applied_at else "",
        )
        ws_tracker.cell(row=row_idx, column=7, value="")
        ws_tracker.cell(row=row_idx, column=8, value=app.notes)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def export_jobs_to_csv(db: Session) -> str:
    """Export jobs to CSV format."""
    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow([
        "Rank", "Job Title", "Company", "Location", "Remote Type",
        "Posted", "Job URL", "Match Score", "Interview Probability",
        "Recommendation", "Status",
    ])

    jobs = (
        db.query(Job)
        .filter(Job.match_score.isnot(None))
        .order_by(Job.match_score.desc())
        .all()
    )

    for rank, job in enumerate(jobs, 1):
        writer.writerow([
            rank,
            job.title,
            job.company,
            job.location,
            job.remote_type,
            job.posted_at.strftime("%Y-%m-%d") if job.posted_at else "",
            job.url,
            job.match_score,
            job.interview_probability or "",
            job.recommendation or "",
            job.status,
        ])

    return output.getvalue()


def _get_matching_skills(job: Job) -> list[str]:
    skills = []
    thresholds = [
        ("AWS", job.aws_match),
        ("Kubernetes", job.kubernetes_match),
        ("Terraform", job.terraform_match),
        ("CI/CD", job.cicd_match),
        ("DevSecOps", job.devsecops_match),
        ("Python", job.python_match),
        ("GitOps", job.gitops_match),
    ]
    for name, score in thresholds:
        if score and score >= 70:
            skills.append(f"{name} ({score:.0f}%)")
    return skills


def _get_strongest_evidence(job: Job) -> str:
    scores = {
        "AWS": job.aws_match or 0,
        "Kubernetes": job.kubernetes_match or 0,
        "Terraform": job.terraform_match or 0,
        "CI/CD": job.cicd_match or 0,
        "DevSecOps": job.devsecops_match or 0,
    }
    top = sorted(scores.items(), key=lambda x: x[1], reverse=True)[:3]
    return ", ".join(f"{k}: {v:.0f}%" for k, v in top if v > 0)


def _get_priority(job: Job | None) -> str:
    if not job:
        return ""
    if job.match_score and job.match_score >= 90:
        return "HIGH"
    elif job.match_score and job.match_score >= 85:
        return "MEDIUM"
    return "LOW"


def _fmt_list(json_str: str) -> str:
    try:
        items = json.loads(json_str) if json_str else []
        return ", ".join(items) if items else ""
    except (json.JSONDecodeError, TypeError):
        return str(json_str)
